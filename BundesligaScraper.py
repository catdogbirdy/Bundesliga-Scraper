import bs4
from requests import get
from urllib.request import urlopen as ureq
from urllib.request import urlretrieve as uret
from bs4 import BeautifulSoup as soup
from openpyxl import *
from tkinter import *
import time
import re
from selenium import webdriver
import os
import datetime

def test():

    workbook = 'testbook1.xlsx'

    playerdb = load_workbook(workbook)

    sheet = playerdb.get_sheet_by_name('Injury_Off')

    #get links
    textfile = open("linklist.txt")
    print('File Opened')
    lines = textfile.read().split("\n")

    #List of Daylight Saving dates
    ds2018 = datetime.datetime.strptime('Mar 25, 2018', "%b %d, %Y")
    ds2017 = datetime.datetime.strptime('Mar 26, 2017', "%b %d, %Y")
    ds2016 = datetime.datetime.strptime('Mar 27, 2016', "%b %d, %Y")
    ds2015 = datetime.datetime.strptime('Mar 29, 2015', "%b %d, %Y")
    ds2014 = datetime.datetime.strptime('Mar 30, 2014', "%b %d, %Y")
    ds2013 = datetime.datetime.strptime('Mar 31, 2013', "%b %d, %Y")
    ds2012 = datetime.datetime.strptime('Mar 25, 2012', "%b %d, %Y")
    ds2011 = datetime.datetime.strptime('Mar 27, 2011', "%b %d, %Y")
    ds2010 = datetime.datetime.strptime('Mar 28, 2010', "%b %d, %Y")

    doff2018 = datetime.datetime.strptime('Oct 28, 2018', "%b %d, %Y")
    doff2017 = datetime.datetime.strptime('Oct 29, 2017', "%b %d, %Y")
    doff2016 = datetime.datetime.strptime('Oct 30, 2016', "%b %d, %Y")
    doff2015 = datetime.datetime.strptime('Oct 25, 2015', "%b %d, %Y")
    doff2014 = datetime.datetime.strptime('Oct 26, 2014', "%b %d, %Y")
    doff2013 = datetime.datetime.strptime('Oct 27, 2013', "%b %d, %Y")
    doff2012 = datetime.datetime.strptime('Oct 28, 2012', "%b %d, %Y")
    doff2011 = datetime.datetime.strptime('Oct 30, 2011', "%b %d, %Y")
    doff2010 = datetime.datetime.strptime('Oct 31, 2010', "%b %d, %Y")

    linkslist = []
    for lines in lines:
        if lines[:3] != 'htt':
            continue

        linkslist.append(lines)

    print(linkslist)

    #loop through all link
    for links in linkslist:

        starttime = time.time()

        print(str(linkslist.index(links)) + ' out of ' + str(len(linkslist)))

        global maxrow
        maxrow = sheet.max_row

        noreport = 0
        datecat = 1

        myurl = links

        browser = webdriver.Chrome("/Users/Qixiang/Dropbox/ICS/venv/chromedriver")
        browser.get(myurl)
        time.sleep(0.50)
        pagehtml = browser.page_source

        # parsing as html
        pagesoup = soup(pagehtml, "html.parser")

        # FOR TESTING
        tabs = pagesoup.find_all("div", {"class", "subnavi_box"})
        statistics = re.findall(r'Statistics..\n.*Statistics', str(tabs))
        statistics = re.findall(r'/.*[0-9]', str(statistics))

        statspagelink = 'https://www.transfermarkt.co.uk' + str(statistics[0])

        print(statspagelink)

        #MATCH DETAILS
        gamescore = pagesoup.find("div", {"class", "sb-endstand"}).text
        gamescore = re.findall(r'[0-9{1,}]:[0-9{1,}]', gamescore)
        gamescore = gamescore[0]
        gamescore = gamescore.split(":")
        homescore = gamescore[0]
        awayscore = gamescore[1]

        totalconceded = int(homescore) + int(awayscore)

        date = pagesoup.find("div", {"class", "sb-spieldaten"})
        date = date.find("p", {"class", "sb-datum hide-for-small"})
        date = re.sub("<.*?>", "", str(date))
        dateplayed = re.findall(r'[a-zA-Z]{3} [0-9]{1,}, [0-9]{4}', date)
        dateplayed = str(dateplayed[0])
        # print(dateplayed)
        dateplayed = datetime.datetime.strptime(dateplayed, "%b %d, %Y")
        # print(dateplayed.year)


        # For calculating days from Daylight Savings Day
        def calculatedate():
            global category
            if datecat != 0:
                if 8 <= int(dsONdiff[0]) <= 14:
                    print('Category A - Pre')
                    category = '0'

                if -8 < int(dsONdiff[0]) < 8:
                    print('Category B - Week of ')
                    category = '1'

                if -14 <= int(dsONdiff[0]) <= -8:
                    print('Category C - Post')
                    category = '2'

                if int(dsONdiff[0]) > 14:
                    print('Category D - DNC')
                    category = '3'

                if int(dsONdiff[0]) < -14:
                    print('Category D - DNC')
                    category = '3'

            if datecat == 0:
                category = '1'


# Calculating Offset data
        if int(dateplayed.year) == int(doff2010.year):
            print(doff2010 - dateplayed)

            if doff2010 != dateplayed:
                dsONdiff = doff2010 - dateplayed

                dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
                dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
                print(dsONdiff)

                calculatedate()

            if doff2010 == dateplayed:
                dsONdiff = 0
                datecat = 0
                calculatedate()

        if int(dateplayed.year) == int(doff2011.year):
            print(doff2011 - dateplayed)

            if doff2011 != dateplayed:
                dsONdiff = doff2011 - dateplayed

                dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
                dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
                print(dsONdiff)

                calculatedate()

            if doff2011 == dateplayed:
                dsONdiff = 0
                datecat = 0
                calculatedate()


        if int(dateplayed.year) == int(doff2012.year):
            if doff2012 != dateplayed:
                dsONdiff = doff2012 - dateplayed

                dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
                dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
                print(dsONdiff)

                calculatedate()

            if doff2012 == dateplayed:
                dsONdiff = 0
                datecat = 0
                calculatedate()


        if int(dateplayed.year) == int(doff2013.year):
            if doff2013 != dateplayed:
                dsONdiff = doff2013 - dateplayed

                dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
                dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
                print(dsONdiff)

                calculatedate()

            if doff2013 == dateplayed:
                dsONdiff = 0
                datecat = 0
                calculatedate()


        if int(dateplayed.year) == int(doff2014.year):
            if doff2014 != dateplayed:
                dsONdiff = doff2014 - dateplayed

                dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
                dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
                print(dsONdiff)

                calculatedate()

            if doff2014 == dateplayed:
                dsONdiff = 0
                datecat = 0
                calculatedate()


        if int(dateplayed.year) == int(doff2015.year):
            if doff2015 != dateplayed:
                dsONdiff = doff2015 - dateplayed

                dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
                dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
                print(dsONdiff)

                calculatedate()

            if doff2015 == dateplayed:
                dsONdiff = 0
                datecat = 0
                calculatedate()


        if int(dateplayed.year) == int(doff2016.year):
            if doff2016 != dateplayed:
                dsONdiff = doff2016 - dateplayed

                dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
                dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
                print(dsONdiff)

                calculatedate()

            if doff2016 == dateplayed:
                dsONdiff = 0
                datecat = 0
                calculatedate()


        if int(dateplayed.year) == int(doff2017.year):
            if doff2017 != dateplayed:
                dsONdiff = doff2017 - dateplayed

                dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
                dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
                print(dsONdiff)

                calculatedate()

            if doff2017 == dateplayed:
                dsONdiff = 0
                datecat = 0
                calculatedate()

        if int(dateplayed.year) == int(doff2018.year):
            if doff2018 != dateplayed:
                dsONdiff = doff2018 - dateplayed

                dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
                dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
                print(dsONdiff)

                calculatedate()

            if doff2018 == dateplayed:
                dsONdiff = 0
                datecat = 0
                calculatedate()


# EXTRACTING DLS ONSET DATA
#         if int(dateplayed.year) == int(ds2010.year):
#             print(ds2010 - dateplayed)
#
#             if ds2010 != dateplayed:
#                 dsONdiff = ds2010 - dateplayed
#
#                 dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
#                 dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
#                 print(dsONdiff)
#
#                 calculatedate()
#
#             if ds2010 == dateplayed:
#                 dsONdiff = 0
#                 datecat = 0
#                 calculatedate()
#
#         if int(dateplayed.year) == int(ds2011.year):
#             print(ds2011 - dateplayed)
#
#             if ds2011 != dateplayed:
#                 dsONdiff = ds2011 - dateplayed
#
#                 dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
#                 dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
#                 print(dsONdiff)
#
#                 calculatedate()
#
#             if ds2011 == dateplayed:
#                 dsONdiff = 0
#                 datecat = 0
#                 calculatedate()
#
#
#         if int(dateplayed.year) == int(ds2012.year):
#             if ds2012 != dateplayed:
#                 dsONdiff = ds2012 - dateplayed
#
#                 dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
#                 dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
#                 print(dsONdiff)
#
#                 calculatedate()
#
#             if ds2012 == dateplayed:
#                 dsONdiff = 0
#                 datecat = 0
#                 calculatedate()
#
#
#         if int(dateplayed.year) == int(ds2013.year):
#             if ds2013 != dateplayed:
#                 dsONdiff = ds2013 - dateplayed
#
#                 dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
#                 dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
#                 print(dsONdiff)
#
#                 calculatedate()
#
#             if ds2013 == dateplayed:
#                 dsONdiff = 0
#                 datecat = 0
#                 calculatedate()
#
#
#         if int(dateplayed.year) == int(ds2014.year):
#             if ds2014 != dateplayed:
#                 dsONdiff = ds2014 - dateplayed
#
#                 dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
#                 dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
#                 print(dsONdiff)
#
#                 calculatedate()
#
#             if ds2014 == dateplayed:
#                 dsONdiff = 0
#                 datecat = 0
#                 calculatedate()
#
#
#         if int(dateplayed.year) == int(ds2015.year):
#             if ds2015 != dateplayed:
#                 dsONdiff = ds2015 - dateplayed
#
#                 dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
#                 dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
#                 print(dsONdiff)
#
#                 calculatedate()
#
#             if ds2015 == dateplayed:
#                 dsONdiff = 0
#                 datecat = 0
#                 calculatedate()
#
#
#         if int(dateplayed.year) == int(ds2016.year):
#             if ds2016 != dateplayed:
#                 dsONdiff = ds2016 - dateplayed
#
#                 dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
#                 dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
#                 print(dsONdiff)
#
#                 calculatedate()
#
#             if ds2016 == dateplayed:
#                 dsONdiff = 0
#                 datecat = 0
#                 calculatedate()
#
#
#         if int(dateplayed.year) == int(ds2017.year):
#             if ds2017 != dateplayed:
#                 dsONdiff = ds2017 - dateplayed
#
#                 dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
#                 dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
#                 print(dsONdiff)
#
#                 calculatedate()
#
#             if ds2017 == dateplayed:
#                 dsONdiff = 0
#                 datecat = 0
#                 calculatedate()
#
#         if int(dateplayed.year) == int(ds2018.year):
#             if ds2018 != dateplayed:
#                 dsONdiff = ds2018 - dateplayed
#
#                 dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
#                 dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
#                 print(dsONdiff)
#
#                 calculatedate()
#
#             if ds2018 == dateplayed:
#                 dsONdiff = 0
#                 datecat = 0
#                 calculatedate()


        #Extract time that match was played
        timeplayed = re.findall(r'[0-9]+:[0-9]{2}', date)
        timeplayed = timeplayed[0] + ' PM'

        league = pagesoup.find("div", {"class", "spielername-profil"}).text
        league = re.sub("\n", "", league)
        league = str(league)
        print(league)

        #INJURY DETAILS
        matchdetails = pagesoup.find_all("div", {"class", "sb-ereignisse"})
        breakdown = re.sub("<.*?>", "", str(matchdetails))

        #For fouls
        yellowcards = re.findall('Yellow card', breakdown)
        redcards = re.findall('Red card', breakdown)

        yellowcards = str(len(yellowcards))
        redcards = str(len(redcards))


        #For injuries
        injuries = re.findall('Injury', breakdown)
        injuries = str(len(injuries))
        print(injuries)

        #not reported findings
        notreported = re.findall('Not reported', breakdown)
        if len(notreported) > 0:
            noreport = 1

        print('Date Played: ' + str(dateplayed))
        print('Time Played: ' + timeplayed)
        print('Amount of Injuries: ' + injuries)
        print('Yellow Cards: ' + yellowcards)
        print('Red Cards: ' + redcards)
        print('Home Team Score: ' + homescore)
        print('Away Team Score: ' + awayscore)
        print('Total Goals Conceded: ' + str(totalconceded))


        ## This portion writes to excel, comment out when testing

        sheet.cell(row=maxrow + 1, column=1).value = dateplayed
        sheet.cell(row=maxrow + 1, column=2).value = timeplayed
        sheet.cell(row=maxrow + 1, column=3).value = injuries
        sheet.cell(row=maxrow + 1, column=4).value = league
        sheet.cell(row=maxrow + 1, column=5).value = yellowcards
        sheet.cell(row=maxrow + 1, column=6).value = redcards
        sheet.cell(row=maxrow + 1, column=7).value = links
        sheet.cell(row=maxrow + 1, column=9).value = homescore
        sheet.cell(row=maxrow + 1, column=10).value = awayscore
        sheet.cell(row=maxrow + 1, column=11).value = str(totalconceded)
        sheet.cell(row=maxrow + 1, column=12).value = category

        if datecat != 0:
            sheet.cell(row=maxrow + 1, column=13).value = str(dsONdiff[0])

        if datecat == 0:
            sheet.cell(row=maxrow + 1, column=13).value = '0'
            print('DAY OF DLS')





        if noreport == 1:
            sheet.cell(row=maxrow + 1, column=8).value = '1'
            print('Missing Data: YES')

        if noreport != 1:
            sheet.cell(row=maxrow + 1, column=8).value = '0'
            print('Missing Data: NO')


        noreport = 0
        datecat = 1

        browser.close()

        # Open new browser to get match statistics
        browser = webdriver.Chrome("/Users/Qixiang/Dropbox/ICS/venv/chromedriver")
        myurl = statspagelink
        browser.get(myurl)
        time.sleep(0.50)
        pagehtml = browser.page_source

        # parsing as html
        pagesoup = soup(pagehtml, "html.parser")

        # Get match statistics
        matchstatistics = pagesoup.find_all("div", {"class", "sb-statistik"})
        matchstatistics = re.findall(r'>[0-9]{1,}<', str(matchstatistics))
        matchstatsclean = []
        for i in matchstatistics:
            i = re.sub("[><]", "", i)
            print(i)
            matchstatsclean.append(i)
        print(matchstatsclean)
        print(len(matchstatistics))

        counting = 0
        missingdata = matchstatsclean.count('0')

        for i in matchstatsclean:
            if len(matchstatsclean) == 0:
                break
            if missingdata == 14:
                break

            counting += 1
            sheet.cell(row = maxrow + 1, column = 13+counting).value = i

        browser.close()
        playerdb.save(workbook)


        endtime = time.time()
        onelooptime = round(endtime-starttime,2)



        print('This loop took: ' + str(onelooptime) + ' secs.')

        print('//////////////////////////////////////////////')
